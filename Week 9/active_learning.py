from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
from scipy.optimize import curve_fit

print("=== ACTIVE LEARNING FILE IS RUNNING ===")
print("File:", __file__)

# Load Excel file
file_path = "Week 9/Pipe_Material_Training_Data.xlsx"
workbook = load_workbook(file_path)
sheet = workbook.active

# Weibull CDF
def cumulative_density_function(x, c, b, a):
    return 1 - c * np.exp(-((x / b)**a))

def plot_cdf_curve_fit(table, material_type, line_color):
    X_train = []
    Y_train = []

    for row in table.iter_rows(values_only=True):
        mat, life, surv = row
        if mat == "Material Type":
            continue

        if mat == material_type:
            # Skip invalid rows
            if float(life) <= 0:
                continue

            # Convert survival% -> CDF
            cdf_value = 1 - float(surv)/100

            X_train.append(float(life))
            Y_train.append(cdf_value)

    X_train = np.array(X_train)
    Y_train = np.array(Y_train)

    # Fix Copper (all values identical → cannot fit curve)
    if np.all(Y_train == Y_train[0]):
        Y_train = Y_train + 0.001*np.arange(len(Y_train))  # tiny noise so the fit works

    # Fit Weibull model
    coeffs, _ = curve_fit(
        cumulative_density_function,
        X_train,
        Y_train,
        p0=[1, 50, 2],
        maxfev=20000
    )

    c, b, a = coeffs
    print(f"{material_type}: c={c:.4f}, b={b:.4f}, a={a:.4f}")

    # Smooth curve
    x_vals = np.linspace(1, max(X_train)*1.2, 300)
    y_vals = cumulative_density_function(x_vals, c, b, a)

    # Convert CDF → survival% for plotting
    survival_vals = (1 - y_vals) * 100

    plt.plot(x_vals, survival_vals, color=line_color, label=material_type)


plot_cdf_curve_fit(sheet, "Cast Iron", "red")
plot_cdf_curve_fit(sheet, "Ductile Iron", "blue")
plot_cdf_curve_fit(sheet, "Galvanized Iron", "green")
plot_cdf_curve_fit(sheet, "Copper", "brown")

plt.legend()
plt.xlabel("Life Expectancy (y)")
plt.ylabel("Survival Probability (%)")
plt.title("Pipe Material Survival CDF")
plt.grid(True)
plt.show()

# TASK 2 ---------------------------------------------------------
from collections import namedtuple
from datetime import datetime
import math
import csv

def weibull_survival(age, c, b, a):
    return c * math.exp(-((age / b) ** a)) * 100   # convert to %

coefficients = {
    "Cast Iron":        (1.0236, 91.2607, 1.6987),
    "Ductile Iron":     (1.0434, 66.0815, 1.6726),
    "Galvanized Iron":  (1.2914, 25.9335, 1.4309),
    "Copper":           (1.0903, 44.1999, 1.6471)
}

WaterMain = namedtuple(
    "WaterMain",
    ["MainType", "Diameter", "InstallDate", "Material", "Age", "Survival_Probability"]
)

water_mains_table = []

csv_path = "Week 9/Water_Mains.csv"

with open(csv_path, newline='', encoding="windows-1252") as csvfile:
    reader = csv.DictReader(csvfile)
    current_year = 2025

    for row in reader:

        # Fix BOM column name
        if "ï»¿MainType" in row:
            row["MainType"] = row.pop("ï»¿MainType")

        # Parse date
        install_date = datetime.strptime(row["InstallDate"], "%m/%d/%Y %H:%M")
        material = row["Material"]

        # Age
        age = current_year - install_date.year

        # Survival probability
        if material in coefficients:
            c, b, a = coefficients[material]
            survival = weibull_survival(age, c, b, a)
            if survival > 100:
                survival = 100
        else:
            survival = None

        # Create record
        wm = WaterMain(
            MainType=row["MainType"],
            Diameter=row["Diameter"],
            InstallDate=install_date,
            Material=material,
            Age=age,
            Survival_Probability=round(survival, 3) if survival is not None else None
        )

        water_mains_table.append(wm)   # <-- CORRECTLY INDENTED INSIDE LOOP

# Print sample rows
for row in water_mains_table[:5]:
    print(row)



# TASK 3 ---------------------------------------------------------
import matplotlib.pyplot as plt

# Extract survival probabilities from Task 2 results
survival_values = [
    row.Survival_Probability
    for row in water_mains_table
    if row.Survival_Probability is not None
]

# Define threshold for vulnerability
threshold = 70

# Identify pipes below the threshold
low_survival = [v for v in survival_values if v < threshold]

percent_low = (len(low_survival) / len(survival_values)) * 100

# Plot histogram
plt.figure(figsize=(8, 5))
plt.hist(survival_values, bins=15, color='skyblue', edgecolor='black')
plt.axvline(threshold, color='red', linestyle='--', label=f'{threshold}% threshold')

plt.title('Distribution of Pipe Survival Probabilities – Mercator Water')
plt.xlabel('Survival Probability (%)')
plt.ylabel('Number of Pipes')
plt.legend()
plt.grid(axis='y', alpha=0.3)
plt.show()

print(f"Pipes below {threshold}% survival: {len(low_survival)} "
      f"({percent_low:.1f}% of total)")
